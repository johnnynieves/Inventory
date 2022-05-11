#!/usr/bin/env python3

from os import system
from pprint import pprint
import openpyxl as xl
from openpyxl.styles import PatternFill
from tqdm import tqdm


def doinventory(inventory_name, online_systems):
    inventory = get_inventory(inventory_name)
    check = 0
    serialsfound = []
    serialsnotfound = []
    for line in inventory:
        line = line.split(",")
        if line[3]:
            check = check + 1
            checkingNone = (checkonlinesystems(line[3].strip(), online_systems))
            if checkingNone == None:
                serialsnotfound.append(line[3])
            else:
                serialsfound.append(checkingNone)
    print("FOUND:")
    print('+'*80)
    print(serialsfound)
    print('+'*80)
    print()
    print("NOT FOUND:")
    print('-'*80)
    print(serialsnotfound)
    print('-'*80)
    print()
    print('*'*80)
    print(f'Serial numbers in Inventory list: {check}')
    print(f'Serial numbers FOUND against PDQ Inventry list: {len(serialsfound)}')
    print(f'Serial numbers NOT FOUND against PDQ Inventry list: {check - len(serialsfound)}')
    print('*'*80)
    print()
    print('Writing found and not found serial numbers to CSV file...')

    with open("serialsnotfound.csv", "w") as f:
        for item in serialsnotfound:
            f.write(f'{item}\n')
    with open("serialsfound.csv", "w") as f:
        for item in serialsfound:
            f.write(f'{item}\n')
    print('*'*80)
    return serialsfound


def checkonlinesystems(check, online_systems):
    
    onlinesys =  get_online_systems(online_systems)
    serialsfound = []
    serialsnotfound = []
    found = 0
    
    for serial in onlinesys:
        serial = serial.strip().split(",")
        
        if serial[3] != check:
            serialsnotfound.append(check)
        elif serial[3] == check:
            found = found + 1
            return check


def get_inventory(inventory):
    copy_file = create_copy_of_inventory(f'{inventory}')
    with open(f"{inventory}.csv", "r") as f:
        lines = f.readlines()
        # lines =  lines.split(",")
    return lines


def get_online_systems(online_systems):
    # online_systems = input('Online Systems file name: ')
    with open(f'{online_systems}.csv', "r") as f:
        lines = f.readlines()
        return lines
        
def colorcode(found,inventory):
    file = f'{inventory}_new.xlsx'
    print('_'*80)
    wb = xl.load_workbook(f'{file}')
    ws = wb["inventory"]
    fill_green = PatternFill(patternType="solid", fgColor="00FF00")
    # fill_red = PatternFill(patternType="solid", fgColor="FF0000")
    pbar =  tqdm(range(len(found)),desc='Color coding in progress: ')
    for item in found:
        # print(f'Looking for {item}')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                # print(f'Checking rows and cols for item against {cell.value}')
                if cell.value == item:
                    # print(f'Found {item} : cell {cell.value}')
                    cell.fill = fill_green
                    # print('Attempting to colorcode cell Green')
                    wb.save(file)
        pbar.update()
    pbar.close()
    # sorted_inventory(file)
    print('Finished Inventory')
    print(f'Refer to colorcoded excel file {file} on your local directory\nIt shoulb be the same place your initial inventory file was placed.')


# def sorted_inventory(file):
#     pass
#     wb = xl.load_workbook(file)
#     ws = wb["Finished_Inventory"]
#     ws.sort_values(by="Serial Number", cellColor)
#     wb.save(file)
#     print('Sorted Inventory')
#     print('_'*80)



def create_copy_of_inventory(file):
    wb = xl.load_workbook(f'{file}.xlsx')
    ws = wb["inventory"]
    file_new = f'{file}_new.xlsx'
    wb.save(file_new)
    
    

def main():
    try:
        system("clear")
        print('*'*80)
        print('''
        Please follow these Instructions:

        1. Please enter only the name of the inventory file when ask to do so, no extension.
        2. You should have the inventory file in both .csv and .xlsx format.
        3. Your descriptions in each cell should not contain "," fix the "," before you
        Save the .xlsx file as a .csv file.
        4. Ensure the worksheet tab within the .xlsx inventory file is named "inventory".
        5. Place the online_systems file in the same directory as the inventory file.
        It should be in .csv format if you got it from PDQ Inventory.
        6. If you do not follow these instructions an error will occur until you have meet
        the above requirements. 
        ''')
        print('*'*80+'\n')
        input('Press Enter to continue...\n')
        inventory = input('Inventory file name: ')
        online_systems = input('Online Systems file name: ')
        print('*'*80)
        serialsfound = doinventory(inventory, online_systems)
        colorcode(serialsfound,inventory)
    except Exception as e:
        system('clear')
        print(e)
        main()

if __name__ == "__main__":
    main()
